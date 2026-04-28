#!/usr/bin/env python3
"""Parse Chase credit card statement PDFs and extract transactions.

Usage:
    python3 parse_chase_statements.py <pdf> [<pdf> ...]
    python3 parse_chase_statements.py --out transactions.xlsx ~/Downloads/Statements*.pdf

Output columns: file, date, merchant, amount, type, owner
    - date    : ISO date; year inferred from statement Opening/Closing Date
    - merchant: merchant name / transaction description
    - amount  : float; positive = purchase, negative = payment/credit
    - type    : PAYMENT / PURCHASE / CASH ADVANCE / FEE / INTEREST / BALANCE TRANSFER
    - owner   : "c" for Cavan-only purchases, "b" for shared purchases, blank for payments

Requires: pdfplumber, openpyxl (pip install pdfplumber openpyxl)
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook


PERIOD_RE = re.compile(
    r"Opening/Closing Date\s+"
    r"(\d{2}/\d{2}/\d{2})\s*-\s*(\d{2}/\d{2}/\d{2})"
)

SECTION_NAMES = [
    "PAYMENTS AND OTHER CREDITS",
    "PURCHASES",
    "PURCHASE",
    "CASH ADVANCES",
    "CASH ADVANCE",
    "FEES CHARGED",
    "INTEREST CHARGED",
    "BALANCE TRANSFERS",
]

# pdfplumber sometimes emits bold glyphs as doubled letters
# (e.g. "AACCCCOOUUNNTT AACCTTIIVVIITTYY"). Collapse any run of 2+ same chars
# to a single char for marker detection only.
_DUP_RE = re.compile(r"(.)\1+")


def _collapse_doubles(s: str) -> str:
    return _DUP_RE.sub(r"\1", s)


# Collapsed markers to test against; compare both raw and collapsed inputs.
_ACTIVITY_MARKERS = {"ACCOUNT ACTIVITY", _collapse_doubles("ACCOUNT ACTIVITY")}
_SECTION_MARKERS = {name: name for name in SECTION_NAMES}
_SECTION_MARKERS.update({_collapse_doubles(name): name for name in SECTION_NAMES})

# MM/DD <merchant...> <amount>
TX_RE = re.compile(
    r"^\s*(\d{2}/\d{2})\s+"
    r"(.+?)\s+"
    r"(-?\d[\d,]*\.\d{2})\s*$"
)

SECTION_TO_TYPE = {
    "PAYMENTS AND OTHER CREDITS": "PAYMENT",
    "PURCHASE": "PURCHASE",
    "PURCHASES": "PURCHASE",
    "CASH ADVANCE": "CASH ADVANCE",
    "CASH ADVANCES": "CASH ADVANCE",
    "FEES CHARGED": "FEE",
    "INTEREST CHARGED": "INTEREST",
    "BALANCE TRANSFERS": "BALANCE TRANSFER",
}

OWNER_C_PATTERNS = (
    re.compile(r"CHATGPT|OPENAI", re.IGNORECASE),
    re.compile(r"\bALO\b", re.IGNORECASE),
    re.compile(r"MAMMOTH", re.IGNORECASE),
    re.compile(r"SOLIDCORE", re.IGNORECASE),
    re.compile(r"CLASSPASS", re.IGNORECASE),
    re.compile(r"HIGHLINE", re.IGNORECASE),
)


@dataclass
class Transaction:
    file: str
    date: date
    merchant: str
    amount: float
    type: str
    owner: str


def _extract_text(pdf_path: Path) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join((page.extract_text() or "") for page in pdf.pages)


def _statement_period(text: str) -> tuple[date, date]:
    m = PERIOD_RE.search(text)
    if not m:
        raise ValueError("Could not find Opening/Closing Date")
    open_d = datetime.strptime(m.group(1), "%m/%d/%y").date()
    close_d = datetime.strptime(m.group(2), "%m/%d/%y").date()
    return open_d, close_d


def _infer_year(mm: int, open_d: date, close_d: date) -> int:
    """Pick the correct calendar year for a MM/DD transaction.

    Statement periods straddle at most two calendar years (e.g. Dec->Jan).
    If the transaction month is >= opening month, use open year; else close year.
    """
    if open_d.year == close_d.year:
        return open_d.year
    return open_d.year if mm >= open_d.month else close_d.year


def parse_chase_statement(pdf_path: Path) -> list[Transaction]:
    text = _extract_text(pdf_path)
    open_d, close_d = _statement_period(text)

    records: list[Transaction] = []
    in_activity = False
    current_section: str | None = None

    for raw_line in text.splitlines():
        stripped = raw_line.strip()
        collapsed = _collapse_doubles(stripped)

        if stripped in _ACTIVITY_MARKERS or collapsed in _ACTIVITY_MARKERS:
            in_activity = True
            continue
        if not in_activity:
            continue

        # End of transaction area: YTD totals / interest tables.
        if (
            re.match(r"^20\d{2} Totals Year-to-Date", stripped)
            or stripped.startswith("INTEREST CHARGES")
            or collapsed.startswith("INTEREST CHARGES")
            or stripped.startswith("Totals Year-to-Date")
        ):
            in_activity = False
            continue

        if stripped in _SECTION_MARKERS:
            current_section = _SECTION_MARKERS[stripped]
            continue
        if collapsed in _SECTION_MARKERS:
            current_section = _SECTION_MARKERS[collapsed]
            continue

        tx_match = TX_RE.match(raw_line)
        if not tx_match:
            continue

        mmdd, merchant, amount_s = tx_match.groups()
        merchant = re.sub(r"\s+", " ", merchant).strip()
        # Some payment lines have a stray leading "& " artifact from the PDF.
        merchant = re.sub(r"^&\s+", "", merchant)

        mm, dd = int(mmdd[:2]), int(mmdd[3:5])
        year = _infer_year(mm, open_d, close_d)
        tx_date = date(year, mm, dd)
        amount = float(amount_s.replace(",", ""))

        tx_type = SECTION_TO_TYPE.get(current_section or "", "")
        owner = _classify_owner(merchant, amount, tx_type)

        records.append(
            Transaction(
                file=pdf_path.name,
                date=tx_date,
                merchant=merchant,
                amount=amount,
                type=tx_type,
                owner=owner,
            )
        )

    return records


def _classify_owner(merchant: str, amount: float, tx_type: str) -> str:
    if amount <= 0 or tx_type == "PAYMENT":
        return ""
    for pattern in OWNER_C_PATTERNS:
        if pattern.search(merchant):
            return "c"
    return "b"


def _summary_path(out_path: Path) -> Path:
    return out_path.with_name(f"{out_path.stem}_b_monthly_summary{out_path.suffix}")


def _monthly_b_summary(rows: list[Transaction]) -> list[dict[str, str | int | float]]:
    monthly: dict[str, dict[str, int | float | str]] = {}
    for row in rows:
        if row.owner != "b":
            continue
        month = row.date.strftime("%Y-%m")
        if month not in monthly:
            monthly[month] = {"month": month, "b_total": 0.0, "b_count": 0}
        monthly[month]["b_total"] += row.amount
        monthly[month]["b_count"] += 1

    return [
        {
            "month": month,
            "b_total": f"{values['b_total']:.2f}",
            "b_count": str(values["b_count"]),
            "c_share": f"{(values['b_total'] * 2 / 3):.2f}",
            "v_share": f"{(values['b_total'] / 3):.2f}",
        }
        for month, values in sorted(monthly.items())
    ]


def _metadata_rows(rows: list[Transaction], source_pdfs: list[Path]) -> list[dict[str, str]]:
    purchase_count = sum(1 for row in rows if row.amount > 0)
    payment_count = sum(1 for row in rows if row.amount <= 0)
    c_count = sum(1 for row in rows if row.owner == "c")
    b_count = sum(1 for row in rows if row.owner == "b")
    return [
        {"field": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
        {"field": "source_files", "value": ", ".join(pdf.name for pdf in source_pdfs)},
        {"field": "statement_file_count", "value": str(len(source_pdfs))},
        {"field": "transaction_row_count", "value": str(len(rows))},
        {"field": "purchase_row_count", "value": str(purchase_count)},
        {"field": "payment_row_count", "value": str(payment_count)},
        {"field": "owner_c_row_count", "value": str(c_count)},
        {"field": "owner_b_row_count", "value": str(b_count)},
        {
            "field": "owner_c_rules",
            "value": "CHATGPT/OPENAI, ALO, MAMMOTH, SOLIDCORE, CLASSPASS, HIGHLINE",
        },
        {"field": "shared_split", "value": "c=2/3, v=1/3"},
    ]


def _rows_to_dicts(rows: list[Transaction]) -> list[dict[str, str | float]]:
    return [
        {
            "file": row.file,
            "date": row.date.isoformat(),
            "merchant": row.merchant,
            "amount": row.amount,
            "type": row.type,
            "owner": row.owner,
        }
        for row in rows
    ]


def _write_csv(rows: list[Transaction], out_path: Path) -> None:
    with out_path.open("w", newline="") as fh:
        writer = csv.DictWriter(
            fh, fieldnames=["file", "date", "merchant", "amount", "type", "owner"]
        )
        writer.writeheader()
        writer.writerows(_rows_to_dicts(rows))


def _write_b_summary_csv(rows: list[Transaction], out_path: Path) -> Path:
    summary_path = _summary_path(out_path)
    summary_rows = _monthly_b_summary(rows)
    with summary_path.open("w", newline="") as fh:
        writer = csv.DictWriter(
            fh, fieldnames=["month", "b_total", "b_count", "c_share", "v_share"]
        )
        writer.writeheader()
        writer.writerows(summary_rows)
    return summary_path


def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        max_len = max(len(value) for value in values) if values else 0
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 60)


def _write_xlsx(rows: list[Transaction], out_path: Path, source_pdfs: list[Path]) -> None:
    workbook = Workbook()

    tx_sheet = workbook.active
    tx_sheet.title = "Transactions"
    tx_fields = ["file", "date", "merchant", "amount", "type", "owner"]
    tx_sheet.append(tx_fields)
    for row in _rows_to_dicts(rows):
        tx_sheet.append([row[field] for field in tx_fields])

    summary_sheet = workbook.create_sheet("Monthly B Summary")
    summary_fields = ["month", "b_total", "b_count", "c_share", "v_share"]
    summary_sheet.append(summary_fields)
    for row in _monthly_b_summary(rows):
        summary_sheet.append([row[field] for field in summary_fields])

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_fields = ["field", "value"]
    metadata_sheet.append(metadata_fields)
    for row in _metadata_rows(rows, source_pdfs):
        metadata_sheet.append([row[field] for field in metadata_fields])

    _autosize_worksheet(tx_sheet)
    _autosize_worksheet(summary_sheet)
    _autosize_worksheet(metadata_sheet)
    workbook.save(out_path)


def _print_table(rows: list[Transaction]) -> None:
    if not rows:
        print("(no transactions)")
        return
    widths = {
        "file": max(len(r.file) for r in rows),
        "date": 10,
        "merchant": min(60, max(len(r.merchant) for r in rows)),
        "amount": 10,
        "type": max(len(r.type) for r in rows),
        "owner": 5,
    }
    header = (
        f"{'file':<{widths['file']}}  "
        f"{'date':<{widths['date']}}  "
        f"{'merchant':<{widths['merchant']}}  "
        f"{'amount':>{widths['amount']}}  "
        f"{'type':<{widths['type']}}  "
        f"{'owner':<{widths['owner']}}"
    )
    print(header)
    print("-" * len(header))
    for r in rows:
        merchant = r.merchant[: widths["merchant"]]
        print(
            f"{r.file:<{widths['file']}}  "
            f"{r.date.isoformat():<{widths['date']}}  "
            f"{merchant:<{widths['merchant']}}  "
            f"{r.amount:>{widths['amount']},.2f}  "
            f"{r.type:<{widths['type']}}  "
            f"{r.owner:<{widths['owner']}}"
        )


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(
        description="Extract transactions from Chase credit card statement PDFs."
    )
    p.add_argument("pdfs", nargs="+", type=Path, help="Statement PDF files")
    p.add_argument(
        "--out",
        type=Path,
        help="Write results to .xlsx or .csv. If omitted, prints a table to stdout.",
    )
    args = p.parse_args(argv)

    all_tx: list[Transaction] = []
    for pdf in args.pdfs:
        if not pdf.exists():
            print(f"error: {pdf} does not exist", file=sys.stderr)
            return 1
        all_tx.extend(parse_chase_statement(pdf))

    all_tx.sort(key=lambda r: (r.date, r.file))

    if args.out:
        if args.out.suffix.lower() == ".xlsx":
            _write_xlsx(all_tx, args.out, args.pdfs)
            print(f"Wrote {len(all_tx)} transactions to workbook {args.out}")
            print("Workbook tabs: Transactions, Monthly B Summary, Metadata")
        else:
            _write_csv(all_tx, args.out)
            summary_path = _write_b_summary_csv(all_tx, args.out)
            print(f"Wrote {len(all_tx)} transactions to {args.out}")
            print(f"Wrote monthly shared summary to {summary_path}")
    else:
        _print_table(all_tx)

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
